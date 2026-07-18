from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from pypdf import PdfReader

from app.services.exports import export_excel, export_pdf


def sample_report():
    return {
        "start": datetime(2026, 7, 1),
        "end": datetime(2026, 7, 2, 23, 59, 59),
        "income": 100_000,
        "expense": 75_000,
        "net": 25_000,
        "transaction_count": 2,
        "by_category": {"Makan & Minum": 25_000, "Transportasi": 50_000},
        "top_transaction": {"amount": 50_000, "description": "Bensin", "category": "Transportasi"},
        "budget_status": {},
        "rows": [
            {
                "id": 1,
                "occurred_at": datetime(2026, 7, 1, 8),
                "kind": "expense",
                "amount": 25_000,
                "category": "Makan & Minum",
                "description": "Makan",
                "tags": "",
                "created_at": datetime(2026, 7, 1, 8),
            },
            {
                "id": 2,
                "occurred_at": datetime(2026, 7, 2, 9),
                "kind": "expense",
                "amount": 50_000,
                "category": "Transportasi",
                "description": "Bensin",
                "tags": "",
                "created_at": datetime(2026, 7, 2, 9),
            },
        ],
    }


def test_excel_export_has_expected_sheets_and_safe_text(tmp_path):
    report = sample_report()
    report["rows"][0]["description"] = '=HYPERLINK("http://evil","click")'
    path = export_excel(report, tmp_path / "report.xlsx")
    workbook = load_workbook(path, data_only=False)
    assert set(["Ringkasan", "Transaksi", "Kategori", "Anggaran", "Tren Harian"]).issubset(
        workbook.sheetnames
    )
    assert workbook["Transaksi"]["G2"].value.startswith("'")


def test_pdf_export_is_readable(tmp_path):
    path = export_pdf(sample_report(), tmp_path / "report.pdf", title="Koran Keuangan")
    reader = PdfReader(str(path))
    assert len(reader.pages) == 1
    assert "Koran Keuangan" in (reader.pages[0].extract_text() or "")


def test_excel_export_normalizes_timezone_aware_datetimes(tmp_path):
    report = sample_report()
    aware = ZoneInfo("Asia/Makassar")
    report["rows"][0]["occurred_at"] = datetime(2026, 7, 1, 8, tzinfo=aware)
    report["rows"][0]["created_at"] = datetime(2026, 7, 1, 8, tzinfo=aware)
    report["rows"][0]["updated_at"] = datetime(2026, 7, 1, 8, tzinfo=aware)
    path = export_excel(report, tmp_path / "aware.xlsx")
    workbook = load_workbook(path, data_only=False)
    assert workbook["Transaksi"]["B2"].value.tzinfo is None
    assert workbook["Transaksi"]["J2"].value.tzinfo is None
    assert workbook["Transaksi"]["K2"].value.tzinfo is None
