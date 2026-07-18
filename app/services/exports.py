from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.reports import format_idr


def _safe_excel_text(value: object) -> object:
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    # Excel has no timezone-aware datetime type. Preserve the wall-clock value.
    return value.replace(tzinfo=None) if value.tzinfo is not None else value


def export_excel(report: dict, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    summary = book.active
    summary.title = "Ringkasan"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    summary.append(["KORAN KEUANGAN", None])
    summary.append(["Periode", f"{report['start']:%d/%m/%Y} - {report['end']:%d/%m/%Y}"])
    summary.append(["Pemasukan", report["income"]])
    summary.append(["Pengeluaran", report["expense"]])
    summary.append(["Arus kas bersih", report["net"]])
    summary.append(["Jumlah transaksi", report["transaction_count"]])
    summary.append([])
    summary.append(["Kategori", "Pengeluaran"])
    for cell in summary[8]:
        cell.fill = header_fill
        cell.font = header_font
    for category, amount in report["by_category"].items():
        summary.append([_safe_excel_text(category), amount])
    for row in summary.iter_rows(min_row=3, max_col=2):
        row[1].number_format = '"Rp" #,##0'
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 22
    summary.freeze_panes = "A3"
    if report["by_category"]:
        chart = BarChart()
        chart.title = "Pengeluaran per Kategori"
        data = Reference(summary, min_col=2, min_row=8, max_row=8 + len(report["by_category"]))
        cats = Reference(summary, min_col=1, min_row=9, max_row=8 + len(report["by_category"]))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        summary.add_chart(chart, "D2")

    transactions = book.create_sheet("Transaksi")
    columns = [
        "ID",
        "Tanggal",
        "Waktu",
        "Jenis",
        "Nominal",
        "Kategori",
        "Deskripsi",
        "Tag",
        "Sumber",
        "Dibuat",
        "Diperbarui",
    ]
    transactions.append(columns)
    for cell in transactions[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in report["rows"]:
        occurred: datetime = row["occurred_at"]
        transactions.append(
            [
                row.get("id"),
                _excel_datetime(occurred).date(),
                _excel_datetime(occurred).time(),
                row["kind"],
                row["amount"],
                _safe_excel_text(row.get("category") or "Lainnya"),
                _safe_excel_text(row.get("description", "")),
                _safe_excel_text(row.get("tags", "")),
                _safe_excel_text(row.get("source", "telegram")),
                _excel_datetime(row.get("created_at")),
                _excel_datetime(row.get("updated_at")),
            ]
        )
    transactions.freeze_panes = "A2"
    transactions.auto_filter.ref = transactions.dimensions
    for column in transactions.columns:
        letter = column[0].column_letter
        transactions.column_dimensions[letter].width = min(
            max(max(len(str(c.value or "")) for c in column) + 2, 12), 32
        )
    for cell in transactions["E"][1:]:
        cell.number_format = '"Rp" #,##0'

    categories = book.create_sheet("Kategori")
    categories.append(["Kategori", "Total Pengeluaran"])
    for category, amount in report["by_category"].items():
        categories.append([_safe_excel_text(category), amount])
    for cell in categories[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cell in categories["B"][1:]:
        cell.number_format = '"Rp" #,##0'

    budgets = book.create_sheet("Anggaran")
    budgets.append(["Kategori", "Anggaran", "Terpakai", "Sisa", "Persentase"])
    for category, status in report["budget_status"].items():
        budgets.append(
            [
                category,
                status["budget"],
                status["used"],
                status["remaining"],
                status["percent"] / 100,
            ]
        )
    for cell in budgets[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in budgets.iter_rows(min_row=2, max_col=4):
        for cell in row[1:]:
            cell.number_format = '"Rp" #,##0'
    for cell in budgets["E"][1:]:
        cell.number_format = "0%"

    trend = book.create_sheet("Tren Harian")
    trend.append(["Tanggal", "Pemasukan", "Pengeluaran", "Bersih"])
    daily: dict[object, dict[str, int]] = {}
    for row in report["rows"]:
        key = row["occurred_at"].date()
        daily.setdefault(key, {"income": 0, "expense": 0})[row["kind"]] += row["amount"]
    for day, values in sorted(daily.items()):
        trend.append(
            [day, values["income"], values["expense"], values["income"] - values["expense"]]
        )
    for cell in trend[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in trend.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = '"Rp" #,##0'
    book.save(path)
    return path


def export_pdf(report: dict, path: str | Path, title: str = "Koran Keuangan") -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    document = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "Helvetica-Bold"
    story = [Paragraph(title, styles["Title"]), Spacer(1, 5 * mm)]
    story.append(
        Paragraph(f"Edisi: {report['start']:%d/%m/%Y} - {report['end']:%d/%m/%Y}", styles["Normal"])
    )
    story.append(Spacer(1, 3 * mm))
    summary_rows = [
        ["Pemasukan", format_idr(report["income"])],
        ["Pengeluaran", format_idr(report["expense"])],
        ["Arus kas bersih", format_idr(report["net"])],
        ["Jumlah transaksi", str(report["transaction_count"])],
    ]
    table = Table(summary_rows, colWidths=[55 * mm, 55 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF2F8")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([table, Spacer(1, 5 * mm), Paragraph("Rekap kategori", styles["Heading2"])])
    category_rows = [["Kategori", "Total"]] + [
        [cat, format_idr(amount)] for cat, amount in report["by_category"].items()
    ]
    category_table = Table(category_rows, colWidths=[90 * mm, 40 * mm], repeatRows=1)
    category_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend(
        [category_table, Spacer(1, 5 * mm), Paragraph("Daftar transaksi", styles["Heading2"])]
    )
    rows = [["Tanggal", "Jenis", "Nominal", "Kategori", "Deskripsi"]]
    for row in report["rows"]:
        rows.append(
            [
                f"{row['occurred_at']:%d/%m/%Y %H:%M}",
                row["kind"],
                format_idr(row["amount"]),
                row.get("category", "Lainnya"),
                row.get("description", ""),
            ]
        )
    tx_table = Table(rows, colWidths=[28 * mm, 20 * mm, 28 * mm, 38 * mm, 55 * mm], repeatRows=1)
    tx_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("PADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(tx_table)
    document.build(story)
    return path
